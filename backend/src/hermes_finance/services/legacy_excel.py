"""Extract the known Hermes Finance legacy workbook format without DB writes."""

from __future__ import annotations

from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from decimal import Decimal, DecimalException
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from hermes_finance.domain import PercentageRate, RubleAmount
from hermes_finance.services.legacy_month_mapping import read_legacy_month_mapping


@dataclass(frozen=True, slots=True)
class LegacyExtractionWarning:
    code: str
    sheet_name: str
    row: int | None
    message: str


@dataclass(frozen=True, slots=True)
class LegacyMonthExtraction:
    sheet_name: str
    reporting_month: dict[str, int]
    snapshot_date: str
    salary: dict[str, Any] | None
    deposits: tuple[dict[str, Any], ...]
    stocks: tuple[dict[str, Any], ...]
    bonds: tuple[dict[str, Any], ...]
    gold: tuple[dict[str, Any], ...]
    mandatory_expenses: tuple[dict[str, Any], ...]
    saving_allocations: tuple[dict[str, Any], ...]
    cashback: tuple[dict[str, Any], ...]
    debts_receivable: tuple[dict[str, Any], ...]
    debts_payable: tuple[dict[str, Any], ...]
    goals: tuple[dict[str, Any], ...]
    dividends: tuple[dict[str, Any], ...]
    comments: tuple[dict[str, Any], ...]
    control_totals: tuple[dict[str, Any], ...]


@dataclass(frozen=True, slots=True)
class LegacyWorkbookExtraction:
    source_file: str
    months: tuple[LegacyMonthExtraction, ...]
    warnings: tuple[LegacyExtractionWarning, ...] = field(default_factory=tuple)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def _text(value: Any) -> str | None:
    if value is None:
        return None
    result = str(value).strip()
    return result or None


def _decimal(value: Any) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        result = Decimal(str(value).strip())
    except (DecimalException, AttributeError):
        return None
    return result if result.is_finite() else None


def _money(value: Any) -> int | None:
    decimal = _decimal(value)
    return None if decimal is None else RubleAmount.from_decimal(decimal).kopecks


def _rate(value: Any) -> int | None:
    decimal = _decimal(value)
    return None if decimal is None else PercentageRate.from_decimal(decimal).basis_points


def _quantity(value: Any) -> str | None:
    decimal = _decimal(value)
    return None if decimal is None else format(decimal, "f")


def _date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def _is_formula(cell: Any) -> bool:
    return cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("="))


def _label(value: Any) -> str:
    """Normalize static Russian legacy labels without interpreting user data."""
    return "".join(character for character in str(value or "").casefold() if character.isalnum())


def _find_label_row(ws: Any, expected: str) -> int | None:
    for row in range(1, ws.max_row + 1):
        if _label(ws[f"A{row}"].value) == expected:
            return row
    return None


def _is_instrument_header(ws: Any, row: int) -> bool:
    return (
        _label(ws[f"A{row}"].value) == "название"
        and _label(ws[f"B{row}"].value) == "isin"
        and _label(ws[f"C{row}"].value) == "типсчёта"
    )


def _instrument_rows(ws: Any, section: str, boundaries: tuple[str, ...]) -> range:
    """Return rows inside one known legacy instrument block, excluding summaries."""
    section_row = _find_label_row(ws, section)
    if section_row is None:
        return range(0)
    header_row = section_row + 1
    if not _is_instrument_header(ws, header_row):
        return range(0)
    boundary_rows = [
        row
        for boundary in boundaries
        if (row := _find_label_row(ws, boundary)) is not None and row > header_row
    ]
    return range(header_row + 1, min(boundary_rows, default=ws.max_row + 1))


def _instrument_row(ws: Any, cached: Any, row: int, *, kind: str) -> dict[str, Any] | None:
    name = _text(ws[f"A{row}"].value)
    if not name:
        return None
    result: dict[str, Any] = {
        "source_row": row,
        "name": name,
        "isin": _text(ws[f"B{row}"].value),
        "account": _text(ws[f"C{row}"].value),
    }
    result["quantity"] = _quantity(ws[f"D{row}"].value)
    if kind == "stock":
        result["cost_kopecks"] = _money(ws[f"E{row}"].value)
        result["market_value_kopecks"] = _money(cached[f"H{row}"].value)
    else:
        result.update(
            {
                "balance_price_percent": _decimal_text(ws[f"E{row}"].value),
                "payments": _money(ws[f"F{row}"].value),
                "market_price_percent": _decimal_text(ws[f"G{row}"].value),
                "market_value_kopecks": _money(cached[f"H{row}"].value),
                "monthly_coupon_kopecks": _money(cached[f"I{row}"].value),
                "payments_per_year": _quantity(ws[f"J{row}"].value),
                "annual_coupon_kopecks": _money(cached[f"K{row}"].value),
                "yield_basis_points": _rate(cached[f"L{row}"].value),
                "payment_date": _date(ws[f"M{row}"].value),
                "notes": _text(ws[f"N{row}"].value),
            }
        )
    return result


def _decimal_text(value: Any) -> str | None:
    decimal = _decimal(value)
    return None if decimal is None else format(decimal, "f")


def _simple_rows(
    ws: Any,
    start: int,
    end: int,
    *,
    name_column: str,
    amount_column: str,
    note_column: str | None = None,
) -> tuple[dict[str, Any], ...]:
    rows: list[dict[str, Any]] = []
    for row in range(start, end + 1):
        name = _text(ws[f"{name_column}{row}"].value)
        amount = _money(ws[f"{amount_column}{row}"].value)
        if not name and amount is None:
            continue
        if name and amount is not None:
            item: dict[str, Any] = {"name": name, "amount_kopecks": amount}
            if note_column:
                item["notes"] = _text(ws[f"{note_column}{row}"].value)
            rows.append(item)
    return tuple(rows)


def _comments(ws: Any, start: int) -> tuple[dict[str, Any], ...]:
    rows: list[dict[str, Any]] = []
    for row in range(start, ws.max_row + 1):
        text = _text(ws[f"B{row}"].value)
        if text:
            position = ws[f"A{row}"].value
            rows.append(
                {
                    "position": int(position) if isinstance(position, (int, float)) else None,
                    "text": text,
                }
            )
    return tuple(rows)


def _control_totals(ws: Any, cached: Any, start: int) -> tuple[dict[str, Any], ...]:
    result: list[dict[str, Any]] = []
    for row in range(start, min(ws.max_row, start + 20) + 1):
        label = _text(ws[f"A{row}"].value)
        value = cached[f"B{row}"].value
        if label and value is not None and not isinstance(value, str):
            result.append(
                {
                    "label": label,
                    "unit": "percentage" if "%" in label else "RUB",
                    "value": _decimal_text(value) if "%" in label else _money(value),
                }
            )
    return tuple(result)


def _extract_month(
    ws: Any, cached: Any, entry: Any, warnings: list[LegacyExtractionWarning]
) -> LegacyMonthExtraction:
    salary_value = _money(ws["B9"].value)
    salary = (
        {"name": "salary", "amount_kopecks": salary_value} if salary_value is not None else None
    )
    deposits: list[dict[str, Any]] = []
    for row in range(17, 22):
        name = _text(ws[f"A{row}"].value)
        if not name or _is_formula(ws[f"A{row}"]):
            continue
        deposits.append(
            {
                "name": name,
                "balance_kopecks": _money(ws[f"B{row}"].value),
                "annual_rate_basis_points": _rate(ws[f"C{row}"].value),
                "maturity_date": _date(ws[f"D{row}"].value),
                "expected_monthly_interest_kopecks": _money(cached[f"E{row}"].value),
            }
        )
    bonds = tuple(
        filter(
            None,
            (
                _instrument_row(ws, cached, row, kind="bond")
                for row in _instrument_rows(
                    ws, "облигациисводная", ("акциисводная", "сводкаиитоги")
                )
            ),
        )
    )
    stocks = tuple(
        filter(
            None,
            (
                _instrument_row(ws, cached, row, kind="stock")
                for row in _instrument_rows(ws, "акциисводная", ("сводкаиитоги",))
            ),
        )
    )
    instruments = [*bonds, *stocks]
    seen_isin: set[str] = set()
    for item in instruments:
        isin = item["isin"]
        if not isin:
            warnings.append(
                LegacyExtractionWarning(
                    "empty_isin", ws.title, item["source_row"], "instrument row has no ISIN"
                )
            )
        elif isin in seen_isin:
            warnings.append(
                LegacyExtractionWarning(
                    "duplicate_isin", ws.title, item["source_row"], "ISIN occurs more than once"
                )
            )
        else:
            seen_isin.add(isin)
    gold: list[dict[str, Any]] = []
    for row in range(34, 44):
        name = _text(ws[f"A{row}"].value)
        if name:
            gold.append(
                {
                    "name": name,
                    "grams": _quantity(ws[f"B{row}"].value),
                    "price_per_gram_kopecks": _money(ws[f"C{row}"].value),
                    "current_value_kopecks": _money(cached[f"D{row}"].value),
                    "purchase_price_per_gram_kopecks": _money(ws[f"E{row}"].value),
                    "pnl_kopecks": _money(cached[f"F{row}"].value),
                }
            )
    expenses = _simple_rows(ws, 10, 19, name_column="H", amount_column="I", note_column="J")
    allocations = _simple_rows(ws, 20, 23, name_column="H", amount_column="I")
    cashback = _simple_rows(ws, 25, 27, name_column="H", amount_column="I", note_column="J")
    receivable = _simple_rows(ws, 30, 32, name_column="H", amount_column="I", note_column="J")
    payable = _simple_rows(ws, 34, 38, name_column="H", amount_column="I", note_column="J")
    goals = _simple_rows(ws, 40, 42, name_column="H", amount_column="I")
    dividends: list[dict[str, Any]] = []
    for row in range(30, 32):
        name = _text(ws[f"A{row}"].value)
        if name:
            dividends.append(
                {
                    "period": name,
                    "total_kopecks": _money(ws[f"B{row}"].value),
                    "months": _quantity(ws[f"C{row}"].value),
                    "monthly_kopecks": _money(cached[f"D{row}"].value),
                }
            )
    comments_row = next(
        (
            row
            for row in range(1, ws.max_row + 1)
            if _text(ws[f"A{row}"].value) == "КОММЕНТАРИИ К МЕСЯЦУ"
        ),
        None,
    )
    summary_row = next(
        (row for row in range(1, ws.max_row + 1) if _text(ws[f"A{row}"].value) == "СВОДКА И ИТОГИ"),
        None,
    )
    return LegacyMonthExtraction(
        sheet_name=ws.title,
        reporting_month={"year": entry.reporting_month.year, "month": entry.reporting_month.month},
        snapshot_date=entry.snapshot_date.isoformat(),
        salary=salary,
        deposits=tuple(deposits),
        stocks=stocks,
        bonds=bonds,
        gold=gold,
        mandatory_expenses=expenses,
        saving_allocations=allocations,
        cashback=cashback,
        debts_receivable=receivable,
        debts_payable=payable,
        goals=goals,
        dividends=tuple(dividends),
        comments=_comments(ws, comments_row + 1) if comments_row else (),
        control_totals=_control_totals(ws, cached, summary_row + 1) if summary_row else (),
    )


def extract_legacy_workbook(workbook_path: Path, mapping_path: Path) -> LegacyWorkbookExtraction:
    """Extract the fixed workbook layout into private raw JSON-ready data."""
    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError("legacy extractor supports only .xlsx workbooks")
    mapping = read_legacy_month_mapping(mapping_path)
    workbook: Any | None = None
    cached_workbook: Any | None = None
    try:
        workbook = load_workbook(workbook_path, data_only=False)
        cached_workbook = load_workbook(workbook_path, data_only=True)
    except (InvalidFileException, OSError) as error:
        if workbook is not None:
            workbook.close()
        raise ValueError("legacy workbook could not be opened") from error
    try:
        warnings: list[LegacyExtractionWarning] = []
        months: list[LegacyMonthExtraction] = []
        mapped_sheets = {entry.sheet_name for entry in mapping.mappings if entry.import_flag}
        for entry in mapping.mappings:
            if not entry.import_flag:
                continue
            if entry.sheet_name not in workbook.sheetnames:
                warnings.append(
                    LegacyExtractionWarning(
                        "mapping_sheet_missing", entry.sheet_name, None, "mapped sheet is absent"
                    )
                )
                continue
            months.append(
                _extract_month(
                    workbook[entry.sheet_name], cached_workbook[entry.sheet_name], entry, warnings
                )
            )
        for sheet_name in workbook.sheetnames:
            if (
                sheet_name not in mapped_sheets
                and "_" in sheet_name
                and sheet_name not in {"Шаблон"}
            ):
                warnings.append(
                    LegacyExtractionWarning(
                        "month_sheet_unmapped",
                        sheet_name,
                        None,
                        "month-like sheet is not importable in mapping",
                    )
                )
        return LegacyWorkbookExtraction(workbook_path.name, tuple(months), tuple(warnings))
    finally:
        if workbook is not None:
            workbook.close()
        if cached_workbook is not None:
            cached_workbook.close()
